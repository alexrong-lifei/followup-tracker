
import io
from datetime import datetime, date

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Follow-Up Tracker", layout="wide")

TRACKER_SHEET = "Follow-Up Tracker"
TODAY_SHEET = "Today Must Do"

HEADERS = [
    "Task ID",
    "Task Description",
    "Vendor / Stakeholder",
    "Owner",
    "Current Responsible",
    "Priority",
    "Status",
    "Start Date",
    "Due Date",
    "Last Follow-Up",
    "Next Follow-Up",
    "Days Since Last Follow-Up",
    "Action Today?",
    "Notes",
]

DATE_COLUMNS = ["Start Date", "Due Date", "Last Follow-Up", "Next Follow-Up"]
STATUS_OPTIONS = ["Waiting", "In Progress", "Done"]
PRIORITY_OPTIONS = ["High", "Medium", "Low"]

OVERDUE_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
TODAY_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")


def empty_tracker_df() -> pd.DataFrame:
    return pd.DataFrame(columns=HEADERS)


def is_blank(value) -> bool:
    try:
        return value is None or value == "" or pd.isna(value)
    except Exception:
        return value is None or value == ""


def coerce_date(value):
    if is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return None
        return value.date()
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                pass
        parsed = pd.to_datetime(value, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.date()
    return None


def display_date(value):
    d = coerce_date(value)
    return d.strftime("%d-%b-%Y") if d else ""


def compute_action_today(row, today=None):
    today = today or date.today()
    status = str(row.get("Status", "")).strip().lower()
    if status == "done":
        return ""

    next_follow = coerce_date(row.get("Next Follow-Up"))
    due_date = coerce_date(row.get("Due Date"))
    responsible = str(row.get("Current Responsible", "")).strip().lower()

    needs_action = False
    if next_follow is not None and next_follow <= today:
        needs_action = True
    if due_date is not None and due_date <= today and "alex" in responsible:
        needs_action = True

    return "Yes" if needs_action else ""


def refresh_computed_fields(df: pd.DataFrame, today=None) -> pd.DataFrame:
    today = today or date.today()
    df = df.copy()

    for header in HEADERS:
        if header not in df.columns:
            df[header] = ""

    df = df[HEADERS]

    for col in DATE_COLUMNS:
        df[col] = df[col].apply(coerce_date)

    df["Days Since Last Follow-Up"] = df["Last Follow-Up"].apply(
        lambda v: (today - coerce_date(v)).days if coerce_date(v) is not None else ""
    )
    df["Action Today?"] = df.apply(lambda r: compute_action_today(r, today=today), axis=1)
    return df


def load_tracker_from_workbook(file_bytes: bytes) -> pd.DataFrame:
    bio = io.BytesIO(file_bytes)
    wb = load_workbook(bio)
    if TRACKER_SHEET not in wb.sheetnames:
        return empty_tracker_df()

    ws = wb[TRACKER_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return empty_tracker_df()

    headers = [str(x).strip() if x is not None else "" for x in rows[0]]
    data = rows[1:]
    df = pd.DataFrame(data, columns=headers)

    for header in HEADERS:
        if header not in df.columns:
            df[header] = ""

    return refresh_computed_fields(df)


def make_sample_df():
    return refresh_computed_fields(empty_tracker_df())


def format_tracker_for_export(df: pd.DataFrame) -> pd.DataFrame:
    out = refresh_computed_fields(df.copy())
    for col in DATE_COLUMNS:
        out[col] = out[col].apply(display_date)
    return out


def auto_style_tracker_sheet(ws, raw_df: pd.DataFrame):
    widths = {
        "A": 10, "B": 38, "C": 24, "D": 12, "E": 20, "F": 10, "G": 14,
        "H": 14, "I": 14, "J": 16, "K": 16, "L": 22, "M": 14, "N": 45
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    today = date.today()
    for excel_row_idx in range(2, ws.max_row + 1):
        df_idx = excel_row_idx - 2
        if df_idx >= len(raw_df):
            continue

        row = raw_df.iloc[df_idx]
        status = str(row.get("Status", "")).strip().lower()
        if status == "done":
            continue

        next_follow = coerce_date(row.get("Next Follow-Up"))
        action_today = str(row.get("Action Today?", "")).strip().lower() == "yes"

        if next_follow is not None and next_follow < today:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(excel_row_idx, col_idx).fill = OVERDUE_FILL
        elif action_today:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(excel_row_idx, col_idx).fill = TODAY_FILL


def build_workbook_bytes(df: pd.DataFrame) -> bytes:
    df = refresh_computed_fields(df)
    export_df = format_tracker_for_export(df)

    wb = Workbook()
    ws = wb.active
    ws.title = TRACKER_SHEET
    ws.append(HEADERS)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for row in export_df.itertuples(index=False):
        ws.append(list(row))

    auto_style_tracker_sheet(ws, df)

    today_df = df[
        (df["Action Today?"] == "Yes")
        & (df["Status"].fillna("").str.lower() != "done")
    ].copy()

    ws2 = wb.create_sheet(TODAY_SHEET)
    ws2.append(HEADERS)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    today_export = format_tracker_for_export(today_df)
    for row in today_export.itertuples(index=False):
        ws2.append(list(row))

    auto_style_tracker_sheet(ws2, today_df)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def dataframe_for_display(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in DATE_COLUMNS:
        out[col] = out[col].apply(display_date)
    return out


if "tracker_df" not in st.session_state:
    st.session_state.tracker_df = make_sample_df()

st.title("Follow-Up Tracker Web App")
st.caption("新增 / 编辑任务，自动刷新 Today Must Do，自动标记 overdue，并导出 Excel。")

with st.sidebar:
    st.subheader("导入或新建")
    uploaded = st.file_uploader("上传你现有的 tracker Excel", type=["xlsx"])
    if uploaded is not None:
        st.session_state.tracker_df = load_tracker_from_workbook(uploaded.getvalue())
        st.success("已载入 Excel，并自动刷新计算列。")

    if st.button("新建空白 Tracker"):
        st.session_state.tracker_df = make_sample_df()
        st.success("已新建空白 Tracker。")

df = refresh_computed_fields(st.session_state.tracker_df.copy())
st.session_state.tracker_df = df

tab1, tab2, tab3, tab4 = st.tabs(["Today Must Do", "全部任务", "新增任务", "导出 Excel"])

with tab1:
    today_df = df[
        (df["Action Today?"] == "Yes")
        & (df["Status"].fillna("").str.lower() != "done")
    ].copy()

    overdue_count = int(sum(
        today_df["Next Follow-Up"].apply(
            lambda x: (coerce_date(x) is not None) and (coerce_date(x) < date.today())
        )
    ))

    c1, c2, c3 = st.columns(3)
    c1.metric("Today Must Do", len(today_df))
    c2.metric("Overdue", overdue_count)
    c3.metric("Total Open", int(sum(df["Status"].fillna("").str.lower() != "done")))

    st.dataframe(dataframe_for_display(today_df), use_container_width=True, hide_index=True)

with tab2:
    st.subheader("筛选与编辑")

    left, right = st.columns([1, 3])
    with left:
        filter_status = st.multiselect("Status", STATUS_OPTIONS, default=STATUS_OPTIONS)
        filter_priority = st.multiselect("Priority", PRIORITY_OPTIONS, default=PRIORITY_OPTIONS)
        filter_action_today = st.selectbox("Action Today?", ["All", "Yes", "No"], index=0)
        keyword = st.text_input("关键词搜索", placeholder="task / vendor / notes")

    filtered = df.copy()
    filtered = filtered[filtered["Status"].isin(filter_status)]
    filtered = filtered[filtered["Priority"].isin(filter_priority)]
    if filter_action_today == "Yes":
        filtered = filtered[filtered["Action Today?"] == "Yes"]
    elif filter_action_today == "No":
        filtered = filtered[filtered["Action Today?"] != "Yes"]

    if keyword.strip():
        mask = pd.Series(False, index=filtered.index)
        for col in ["Task ID", "Task Description", "Vendor / Stakeholder", "Current Responsible", "Notes"]:
            mask = mask | filtered[col].fillna("").astype(str).str.contains(keyword, case=False, na=False)
        filtered = filtered[mask]

    st.dataframe(dataframe_for_display(filtered), use_container_width=True, hide_index=True)

    st.markdown("### 编辑单条任务")
    if len(df) == 0:
        st.info("当前没有任务。先去“新增任务”里添加。")
    else:
        task_options = [f"{row['Task ID']} | {row['Task Description']}" for _, row in df.iterrows()]
        selected = st.selectbox("选择要编辑的任务", options=task_options)
        selected_id = selected.split("|")[0].strip()
        idx = df.index[df["Task ID"] == selected_id][0]
        row = df.loc[idx]

        with st.form("edit_task_form"):
            col1, col2 = st.columns(2)
            with col1:
                task_desc = st.text_input("Task Description", value=str(row["Task Description"]))
                vendor = st.text_input("Vendor / Stakeholder", value=str(row["Vendor / Stakeholder"]))
                owner = st.text_input("Owner", value=str(row["Owner"]))
                responsible = st.text_input("Current Responsible", value=str(row["Current Responsible"]))
                priority_index = PRIORITY_OPTIONS.index(row["Priority"]) if row["Priority"] in PRIORITY_OPTIONS else 0
                priority = st.selectbox("Priority", PRIORITY_OPTIONS, index=priority_index)
            with col2:
                status_index = STATUS_OPTIONS.index(row["Status"]) if row["Status"] in STATUS_OPTIONS else 0
                status = st.selectbox("Status", STATUS_OPTIONS, index=status_index)
                start_date = st.date_input("Start Date", value=coerce_date(row["Start Date"]) or None)
                due_date = st.date_input("Due Date", value=coerce_date(row["Due Date"]) or None)
                last_follow = st.date_input("Last Follow-Up", value=coerce_date(row["Last Follow-Up"]) or None)
                next_follow = st.date_input("Next Follow-Up", value=coerce_date(row["Next Follow-Up"]) or None)

            notes = st.text_area("Notes", value=str(row["Notes"]))

            submitted = st.form_submit_button("保存修改")
            if submitted:
                df.loc[idx, "Task Description"] = task_desc
                df.loc[idx, "Vendor / Stakeholder"] = vendor
                df.loc[idx, "Owner"] = owner
                df.loc[idx, "Current Responsible"] = responsible
                df.loc[idx, "Priority"] = priority
                df.loc[idx, "Status"] = status
                df.loc[idx, "Start Date"] = start_date
                df.loc[idx, "Due Date"] = due_date
                df.loc[idx, "Last Follow-Up"] = last_follow
                df.loc[idx, "Next Follow-Up"] = next_follow
                df.loc[idx, "Notes"] = notes
                st.session_state.tracker_df = refresh_computed_fields(df)
                st.success("任务已更新。")

with tab3:
    st.subheader("新增任务")
    with st.form("add_task_form"):
        col1, col2 = st.columns(2)
        with col1:
            task_id = st.text_input("Task ID", placeholder="例如 T025")
            task_desc = st.text_input("Task Description")
            vendor = st.text_input("Vendor / Stakeholder")
            owner = st.text_input("Owner", value="Alex")
            responsible = st.text_input("Current Responsible")
            priority = st.selectbox("Priority", PRIORITY_OPTIONS, index=0)
        with col2:
            status = st.selectbox("Status", STATUS_OPTIONS, index=0, key="add_status")
            start_date = st.date_input("Start Date", value=None, key="add_start")
            due_date = st.date_input("Due Date", value=None, key="add_due")
            last_follow = st.date_input("Last Follow-Up", value=None, key="add_last")
            next_follow = st.date_input("Next Follow-Up", value=None, key="add_next")

        notes = st.text_area("Notes", placeholder="补充背景、承诺时间、风险等")

        add_submitted = st.form_submit_button("新增任务")
        if add_submitted:
            if not task_id.strip():
                st.error("Task ID 不能为空。")
            elif task_id in df["Task ID"].astype(str).tolist():
                st.error("Task ID 已存在。")
            else:
                new_row = {
                    "Task ID": task_id.strip(),
                    "Task Description": task_desc.strip(),
                    "Vendor / Stakeholder": vendor.strip(),
                    "Owner": owner.strip(),
                    "Current Responsible": responsible.strip(),
                    "Priority": priority,
                    "Status": status,
                    "Start Date": start_date,
                    "Due Date": due_date,
                    "Last Follow-Up": last_follow,
                    "Next Follow-Up": next_follow,
                    "Days Since Last Follow-Up": "",
                    "Action Today?": "",
                    "Notes": notes.strip(),
                }
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                st.session_state.tracker_df = refresh_computed_fields(df)
                st.success(f"任务 {task_id} 已新增。")

with tab4:
    st.subheader("导出带格式的 Excel")
    export_bytes = build_workbook_bytes(df)
    st.download_button(
        "下载 Excel",
        data=export_bytes,
        file_name="Follow_Up_Tracker_WebApp_Export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
